from openpyxl import Workbook
from collections import defaultdict

cities = []
types = []

with open("res/city_id.txt", "r", encoding="utf-8") as f:
    ct = f.readline()
    for i in range(0, int(ct)):
        l = f.readline()[:-1]
        cities.append(l.split(","))

with open("res/type_id.txt", "r", encoding="utf-8") as f:
    ct = f.readline()
    for i in range(int(ct)):
        l = f.readline()[:-1]
        arr = l.split(",")
        if arr[0] in ["超级票", "优惠券", "票票堂", "电影优惠券",
                      "周边商品", "旅游演艺", "韩流地带", "电影", "活动"]:
            continue
        types.append(arr)

wb = Workbook()

blue = '20 % - Accent1'
red = '20 % - Accent2'
green = '20 % - Accent3'

blueTitle = '40 % - Accent1'
redTitle = '40 % - Accent2'
greenTitle = '40 % - Accent3'

ws = wb.active

ws.cell(row=1, column=1, value="城市").style = blueTitle
ws.cell(row=1, column=2, value="总计").style = blueTitle
ws.cell(row=1, column=3, value="场馆数").style = blueTitle
i = 4
for type, type_id in types:
    ws.cell(row=1, column=i, value=type).style = redTitle
    ws.cell(row=1, column=i + 1, value="最低价和").style = greenTitle
    ws.cell(row=1, column=i + 2, value="最高价和").style = greenTitle
    ws.cell(row=1, column=i + 3, value="平均最低价").style = greenTitle
    ws.cell(row=1, column=i + 4, value="平均最高价").style = greenTitle
    i += 5

i = 2
for city, city_id in cities:
    print(city)
    ws = wb.create_sheet(city)
    building_count = defaultdict(lambda: 0)
    with open("res/" + city + ".txt", encoding="utf-8") as f:
        lines = f.readlines()
        ws['A1'] = "类型"
        ws['A1'].style = greenTitle
        ws['B1'] = "地点"
        ws['B1'].style = blueTitle
        ws['C1'] = "名称"
        ws['C1'].style = blueTitle
        ws['D1'] = "链接"
        ws['D1'].style = blueTitle
        ws['E1'] = "最低价"
        ws['E1'].style = blueTitle
        ws['F1'] = "最高价"
        ws['F1'].style = blueTitle
        r = 2
        for line in lines:
            cols = line.split(",")
            if len(cols) == 6:
                c = 1
                for col in cols:
                    col = col.strip()
                    color = blue
                    if c == 1:
                        color = green
                    if c >= 5:
                        if col.isnumeric():
                            col = int(col)
                        else:
                            col = None
                    ws.cell(row=r, column=c, value=col).style = color
                    if c == 2:
                        building_count[col] += 1
                    c += 1
                r += 1

    ws = wb.active
    ws.title = "汇总"
    ws.cell(row=i, column=1, value=city).style = blueTitle
    ws.cell(row=i, column=2, value="=COUNTA(" + city + "!$A:$A) - 1").style = blue
    ws.cell(row=i, column=3, value=len(building_count)).style = blue
    r = 4
    for type, _ in types:
        ws.cell(row=i, column=r, value="=COUNTIF(" + city + "!$A:$A, \"" + type + "\")").style = red
        c = ws.cell(row=i, column=r + 1, value="=IFERROR(SUMIF(" + city + "!$A:$A, \""
                                                + type + "\", " + city + "!$E:$E), 0)")
        c.style = green
        c.number_format = "0.00"
        c = ws.cell(row=i, column=r + 2, value="=IFERROR(SUMIF(" + city + "!$A:$A, \""
                                                + type + "\", " + city + "!$F:$F), 0)")
        c.style = green
        c.number_format = "0.00"
        c = ws.cell(row=i, column=r + 3, value="=IFERROR(AVERAGEIF(" + city + "!$A:$A, \""
                                                + type + "\", " + city + "!$E:$E), 0)")
        c.style = green
        c.number_format = "0.00"
        c = ws.cell(row=i, column=r + 4, value="=IFERROR(AVERAGEIF(" + city + "!$A:$A, \""
                                                + type + "\", " + city + "!$F:$F), 0)")
        c.style = green
        c.number_format = "0.00"
        r += 5
    i += 1

wb.save("res/result.xlsx")
